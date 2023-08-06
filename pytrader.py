import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic
from Kiwoom import *
import time
from os import environ
from PyQt5.QtGui import *
import openpyxl as op
import re


form_class = uic.loadUiType("exam_2.ui")[0]


#해상도 고정 함수 추가
def suppress_qt_warning():
    environ["QT_DEVICE_PIXEL_RATIO"] = "0"
    environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    environ["QT_SCREEN_SCALE_FACTORS"] = "1"
    environ["QT_SCALE_FACTOR"] = "1"

class MyWindow(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
            
        self.trade_set = False
        
        self.trade_stocks_done = False

        self.kiwoom = Kiwoom(self) #객체생성
        self.kiwoom.comm_connect() #연결


        self.timer = QTimer(self)
        self.timer.start(1000)
        self.timer.timeout.connect(self.timeout)


        #self.timer3 = QTimer(self)
        #self.timer3.start(30000)
        #self.timer3.timeout.connect(self.timeout3)


        accouns_num = int(self.kiwoom.get_login_info("ACCOUNT_CNT"))
        accounts = self.kiwoom.get_login_info("ACCNO")

        accounts_list = accounts.split(';')[0:accouns_num]
        
        
        self.comboBox.addItems(accounts_list) #콤보박스 1에 계좌목록 추가
        #self.lineEdit.textChanged.connect(self.code_changed)
        #self.pushButton.clicked.connect(self.check_balance)

        self.pushButton_5.clicked.connect(self.kiwoom.get_condition_load)
        self.pushButton_7.clicked.connect(self.change_total_price)

        self.row_count = 0 #tableWidget_3 에서 행 카운트하는용
        self.window_count = 0 #tableWidget_3 화면번호 만드는용
        self.stock_list = [] #주시종목 담은 리스트
        self.stock_ticker_list = [] #주시종목 티커 리스트 
        self.account_number = self.comboBox.currentText() #계좌
        self.take_profit = 0 #익절기준
        
        
        self.pushButton_5.setDisabled(True)
        self.pushButton_2.setDisabled(True)
        self.pushButton_3.setDisabled(True)
        self.pushButton_6.setDisabled(True)
        
        self.lineEdit_9.textChanged.connect(self.price_change)
        self.pushButton_6.clicked.connect(self.kiwoom._condition_search)
        
        self.lineEdit_12.textChanged.connect(self.enable_pushButton_6)
        
        self.gudoc_status = 0
        self.ticker_list = []
        
        
        
        
        #self.lineEdit_8.textChanged.connect(self.profit_percent)# 익절 기준
        
    
    def price_change(self):
        self.pushButton_5.setEnabled(True)

    #종목 ui에 띄우기
    def code_changed(self):
        code = self.lineEdit.text()
        name = self.kiwoom.get_master_code_name(code)
        self.lineEdit_2.setText(name)
        if name != "":
            self.pushButton_3.setEnabled(True)
        else :
            self.pushButton_3.setDisabled(True)
            
        
    #서버연결
    def timeout(self):
        market_start_time = QTime(9, 0, 0)
        current_time = QTime.currentTime()

        if current_time > market_start_time and self.trade_stocks_done is False:
            #self.trade_stocks()
            self.trade_stocks_done = True

        text_time = current_time.toString("hh:mm:ss")
        time_msg = "현재시간: " + text_time


        state = self.kiwoom.get_connect_state()
        if state == 1:
            state_msg = "서버 연결 중"
            
        else:
            state_msg = "서버 미 연결 중"

        self.statusbar.showMessage(state_msg + " | " + time_msg)
        


    """
    #실시간 검색중 TR요청
    def timeout3(self):
        
    
        print("stock_held : " , self.kiwoom.stock_held)    
    
    
        if self.kiwoom.stock_held != []:
            for i in self.kiwoom.stock_held:
                
                time.sleep(0.2)
                
                
                self.kiwoom.set_input_value("종목코드", i)
                self.kiwoom.set_input_value("틱범위", 5)
                self.kiwoom.set_input_value("수정주가구분", 0)
                self.kiwoom.comm_rq_data("opt10080_req", "opt10080", 0, "3000")


    """

    #현재가격저장        
    def present_price(self):
        price = self.kiwoom.price
        self.lineEdit_3.setText(str(price))
        
        
    def change_total_price(self):
        price = format(int(self.lineEdit_11.text()), ",")    
        self.lineEdit_10.setText(str(price))   
     
    """
    #주식 잔고 
    def check_balance(self):
        self.kiwoom.reset_opw00018_output()
        account_number = self.comboBox.currentText()

        self.kiwoom.set_input_value("계좌번호", account_number)
        self.kiwoom.comm_rq_data("opw00018_req", "opw00018", 0, "2000")

        while self.kiwoom.remained_data:
            time.sleep(0.2)
            self.kiwoom.set_input_value("계좌번호", account_number)
            self.kiwoom.comm_rq_data("opw00018_req", "opw00018", 2, "2000")

        # opw00001
        self.kiwoom.set_input_value("계좌번호", account_number)
        self.kiwoom.comm_rq_data("opw00001_req", "opw00001", 0, "2000")

        # balance
        item = QTableWidgetItem(self.kiwoom.d2_deposit)
        item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
        self.tableWidget.setItem(0, 0, item)

        for i in range(1, 6):
            item = QTableWidgetItem(self.kiwoom.opw00018_output['single'][i - 1])
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
            self.tableWidget.setItem(0, i, item)

        self.tableWidget.resizeRowsToContents()

        # Item list
        item_count = len(self.kiwoom.opw00018_output['multi'])
        self.tableWidget_2.setRowCount(item_count)

        for j in range(item_count):
            row = self.kiwoom.opw00018_output['multi'][j]
            for i in range(len(row)):
                item = QTableWidgetItem(row[i])
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                self.tableWidget_2.setItem(j, i, item)

        self.tableWidget_2.resizeRowsToContents()
    """
     #selectFunction 메서드 정의
    def selectFunction(self):
        #filePath 출력하는 부분 초기화
        self.filePath.clear()

        #선택한 엑셀 파일 경로를 받아옴 : 튜플 타입으로 받아오며 0번재 요소가 주소값 string
        path = QFileDialog.getOpenFileName(self, 'Open File', '', 'All File(*);; xlsx File(*.xlsx)')
        #filePath에 현재 읽어온 엑셀 파일 경로를 입력(절대경로)
        self.filePath.setText(path[0])

        #위 절대 경로 활용해 openpyxl workbook 객체 생성
        wb = op.load_workbook(path[0])
        #설정한 workbook의 시트리스트를 읽기
        self.shtlist = wb.sheetnames
        print(self.shtlist)
        
        #시트리스트를 반복문으로 진행
            
        if self.filePath.text() != "":
            self.fileSave.setEnabled(True)
            
            
    #엑셀함수 주시종목에 넣기
    def fileSaveFunction(self):
        filePath = self.filePath.text()
        wb = op.load_workbook(filename = filePath )
        
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        
        print(ws.max_row+1)
        
        for row in range(2,ws.max_row+1):
            row_list_1 = []
            row_dic = {}
            for col in range(1,ws.max_column+1):
                    row_list_1.append(str(ws.cell(row= row, column= col).value))
    
                            
            #빈칸 있을때 for문 넘기기
            if row_list_1[0] == "None":
                    continue
                      
    
    
            if row_list_1[0] != "None":
                row_dic['종목이름'] = self.kiwoom.get_master_code_name(row_list_1[0].zfill(6))
            if row_list_1[1] != "None":
                row_dic['상단선'] = format(int(row_list_1[1]), ",")
            if row_list_1[3] != "None":
                row_dic['하단선'] = format(int(row_list_1[3]), ",")
            if row_list_1[4] != "None":
                row_dic['금액'] = format(int(row_list_1[4]), ",")
            if row_list_1[0] != "None":
                row_dic['티커'] = row_list_1[0].zfill(6)


            

            if row_list_1[2] =="None":
                row_dic['입력개수'] = "2개"
                row_dic['중단선'] = ""
                if row_dic["상단선"] <= row_dic["하단선"]:
                    self.textEdit.append("상단선이 하단선보다 작거나 같음 : "+ row_dic["종목이름"])
                    continue
                
            else:
                row_dic['입력개수'] = "3개"
                row_dic['중단선'] = format(int(row_list_1[2]), ",")
                
                if row_dic["상단선"] <= row_dic["중단선"]:
                    self.textEdit.append("상단선이 중단선보다 작거나 같음 : "+ row_dic["종목이름"])
                    continue
                elif row_dic["중단선"] <= row_dic["하단선"]:
                    self.textEdit.append("중단선이 하단선보다 작거나 같음 : "+ row_dic["종목이름"])
                    continue
                

                
            self.tableWidget_3.setRowCount(self.row_count+1)
            self.tableWidget_3.setColumnCount(8)
            self.tableWidget_3.setItem(self.row_count,0,QTableWidgetItem(row_dic['종목이름']))
            self.tableWidget_3.setItem(self.row_count,1,QTableWidgetItem(row_dic['상단선']))
            self.tableWidget_3.setItem(self.row_count,2,QTableWidgetItem(row_dic['중단선']))
            self.tableWidget_3.setItem(self.row_count,3,QTableWidgetItem(row_dic['하단선']))
            self.tableWidget_3.setItem(self.row_count,4,QTableWidgetItem(row_dic['티커']))
            self.tableWidget_3.setItem(self.row_count,5,QTableWidgetItem(row_dic['금액']))
            self.tableWidget_3.setItem(self.row_count,6,QTableWidgetItem(row_dic['입력개수']))
            self.tableWidget_3.setItem(self.row_count,7,QTableWidgetItem(str(1000+self.window_count)))
            self.row_count+=1
            self.window_count+=1
                    
                    
            self.textEdit.append("종목추가 : "+ row_dic['종목이름'])
                

                    
                
    #port 이름넣기(매도)
    def check_port(self):
        port_name = self.comboBox_2.currentText()
        self.tableWidget_3.setRowCount(self.row_count+1)
        self.tableWidget_3.setColumnCount(1)
        self.tableWidget_3.setItem(self.row_count,0,QTableWidgetItem(port_name))
        self.row_count+=1
        #self.tableWidget_3.resizeRowsToContents()
        self.tableWidget_3.resizeColumnsToContents()
        

    #거래할 종목수 입력시 거래시작 활성화
    def enable_pushButton_6(self):
        self.pushButton_6.setEnabled(True)


    #주시 종목에 설정한 종목 넣기
    def check_stock(self):
        code = self.lineEdit.text()
        name = self.kiwoom.get_master_code_name(code)
        #count = self.comboBox_2.currentText()
        middle_line = self.lineEdit_4.text()
        
        if middle_line == "":
            high = format(int(self.lineEdit_3.text()), ",")
            low = format(int(self.lineEdit_5.text()), ",")
            middle = self.lineEdit_4.text()
            price = format(int(self.lineEdit_9.text()), ",")
            
            
            self.tableWidget_3.setRowCount(self.row_count+1)
            self.tableWidget_3.setColumnCount(8)
            self.tableWidget_3.setItem(self.row_count,0,QTableWidgetItem(name))
            self.tableWidget_3.setItem(self.row_count,1,QTableWidgetItem(high))
            self.tableWidget_3.setItem(self.row_count,2,QTableWidgetItem(str(middle)))
            self.tableWidget_3.setItem(self.row_count,3,QTableWidgetItem(low))
            self.tableWidget_3.setItem(self.row_count,4,QTableWidgetItem(code))
            self.tableWidget_3.setItem(self.row_count,5,QTableWidgetItem(price))
            self.tableWidget_3.setItem(self.row_count,6,QTableWidgetItem("2개"))
            self.tableWidget_3.setItem(self.row_count,7,QTableWidgetItem(str(1000+self.window_count)))
            self.row_count+=1
            self.window_count+=1

            self.textEdit.append("종목추가 : "+ name)    

            self.lineEdit.clear()
            self.lineEdit_3.clear()
            self.lineEdit_4.clear()
            self.lineEdit_5.clear()
            
            
        
        else:

            high = format(int(self.lineEdit_3.text()), ",")
            middle = format(int(self.lineEdit_4.text()), ",")
            low = format(int(self.lineEdit_5.text()), ",")
            price = format(int(self.lineEdit_9.text()), ",")
            
            self.tableWidget_3.setRowCount(self.row_count+1)
            self.tableWidget_3.setColumnCount(8)
            self.tableWidget_3.setItem(self.row_count,0,QTableWidgetItem(name))
            self.tableWidget_3.setItem(self.row_count,1,QTableWidgetItem(high))
            self.tableWidget_3.setItem(self.row_count,2,QTableWidgetItem(middle))
            self.tableWidget_3.setItem(self.row_count,3,QTableWidgetItem(low))
            self.tableWidget_3.setItem(self.row_count,4,QTableWidgetItem(code))
            self.tableWidget_3.setItem(self.row_count,5,QTableWidgetItem(price))
            self.tableWidget_3.setItem(self.row_count,6,QTableWidgetItem("3개"))
            self.tableWidget_3.setItem(self.row_count,7,QTableWidgetItem(str(1000+self.window_count)))
            self.row_count+=1
            self.window_count+=1
            self.textEdit.append("종목추가 : "+ name)

            
            self.lineEdit.clear()
            self.lineEdit_3.clear()
            self.lineEdit_4.clear()
            self.lineEdit_5.clear()
        


            

            
            
    #제거 버튼 눌렀을때 테이블에서 행삭제(매수)      
    def delete_row(self):
        select = self.tableWidget_3.selectedItems()
        for i in select:
            row = i.row()
        
        
            self.tableWidget_3.removeRow(row)
            self.row_count-=1
            #self.plainTextEdit.appendPlainText("선택 종목삭제")
            self.textEdit.append("선택 port 삭제")
         

    #호가 받아오는 함수
    def get_hoga(self, trcode):
        self.kiwoom.set_input_value("종목코드", trcode)
        self.kiwoom.comm_rq_data("opt10004_req", "opt10004", 0, "3000")
        
    #전일종가 받아오는 함수
    def get_last_close(self, trcode):
        self.kiwoom.set_input_value("종목코드", trcode)
        self.kiwoom.comm_rq_data("opt10002_req", "opt10002", 0, "3000")



    #tableWidget_3 에서 값 얻어오기
    def get_label(self):
        init_num = 0
        init_list = []
        while init_num < self.row_count:
            sec_list = []
            for i in range(8):
                sec_list.append(self.tableWidget_3.item(init_num,i).text())
            init_list.append(sec_list)
            init_num += 1
        print(init_list)
        return init_list
                
    
    
    def ready_trade(self, ticker):
        
        name = self.kiwoom.get_master_code_name(ticker)
        
        self.account_number = self.comboBox.currentText()
        
        
        self.kiwoom.dic[name + '_name'] = name
        self.kiwoom.dic[name + '_ticker'] = ticker
        self.kiwoom.dic[name + '_status'] = '초기상태' 
        self.kiwoom.dic[name + '_rebuy'] = 1  
        self.kiwoom.dic[name + '_initial'] = 0 
        self.kiwoom.dic[name + '_buy_count'] = 0 
        self.kiwoom.dic[name + '_sell_price'] = 0 
        self.kiwoom.dic[name + '_rebuy_count'] = 0
        self.kiwoom.dic[name + '_buy_total'] = int(re.sub(r"[^0-9]", "", self.lineEdit_9.text()))

            
        #매도조건 상태 2가지
        self.kiwoom.dic[name + '_sell_status1'] = '초기상태'
        self.kiwoom.dic[name + '_sell_status2'] = '초기상태'
            
        #재매수시 비율
        self.kiwoom.dic[name + '_sec_percent'] = 0
            
        #각 시점 최고가
        self.kiwoom.dic[name + '_init_under'] = 0 
        
        
        #self.plainTextEdit.appendPlainText("거래준비완료 | 종목 :" + name )
        self.textEdit.append("거래준비완료 | 종목 :" + name)
            
        #2%도달 여부(1매수용) (0도달x / 1 도달o )
        self.kiwoom.dic[name + '_reach_two_per'] = 0 

            
        #2%도달 여부(2매수용) (0도달x / 1 도달o )
        self.kiwoom.dic[name + '_reach_two_per2'] = 0 


        

        print("ready_trade")
    


    #거래시작 버튼눌렀을때 주시 종목별 구독
    def trade_start(self,ticker):
        self.account_number = self.comboBox.currentText()

        if ticker not in self.ticker_list:
            self.ticker_list.append(ticker)

        if ticker not in self.kiwoom.dic.values() and ticker != "":
            self.ready_trade(ticker)
        
        print(self.kiwoom.dic)
        #self.plainTextEdit.appendPlainText("-------------거래 시작----------------")
        self.textEdit.append("-------------종목 검색 시작----------------")
        
        if self.gudoc_status == 0:
            self.kiwoom.SetRealReg(1000 +self.window_count , ticker, "20;10", "0")
            self.window_count += 1
            self.gudoc_status = 1
            print('구독성공')
        elif self.gudoc_status != 0 :
            self.kiwoom.SetRealReg(1000 +self.window_count , ticker, "20;10", "1")
            self.window_count += 1
            print('구독성공2')
            
        #self.stock_ticker_list.append(self.stock_list[i][4]) 


    #익절기준 변경점
   # def profit_percent(self):
        #self.take_profit = float(self.lineEdit_8.text())





if __name__ == "__main__":
    
    import sys
    
    suppress_qt_warning()
    
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()


